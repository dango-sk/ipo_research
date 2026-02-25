"""엑셀 리서치 시트 생성 모듈

리브스메드 엑셀 포맷에 맞춰 IPO 리서치 시트를 생성한다.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.settings import REPORTS_DIR

# 스타일
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
NUM_FORMAT_KRW = '#,##0'
NUM_FORMAT_PCT = '0.00%'


def _write_section_header(ws, row: int, title: str, cols: int = 13):
    """섹션 헤더를 작성한다."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
    ws.cell(row=row, column=1, value=title)


def generate_excel(data: dict, company_name: str) -> Path:
    """IPO 리서치 데이터를 엑셀 파일로 생성한다."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = company_name

    # 컬럼 너비 설정
    widths = [14, 16, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # =====================================================================
    # 1. 타이틀
    # =====================================================================
    ws.cell(row=row, column=1, value=company_name).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    row += 2

    # =====================================================================
    # 2. 일정
    # =====================================================================
    _write_section_header(ws, row, "일정")
    row += 1

    offering = data.get("offering", {})
    crawler = data.get("crawler_data", {})

    schedule_items = [
        ("수요예측일", crawler.get("demand_forecast_date", "")),
        ("청약일", offering.get("subscription_date", "") or crawler.get("subscription_date", "")),
        ("납입일", offering.get("payment_date", "")),
        ("상장예정일", crawler.get("listing_date", "")),
        ("주관사", crawler.get("lead_underwriter", "")),
    ]
    for label, val in schedule_items:
        ws.cell(row=row, column=1, value=label).font = HEADER_FONT
        ws.cell(row=row, column=2, value=val)
        row += 1

    row += 1

    # =====================================================================
    # 3. 공모사항
    # =====================================================================
    _write_section_header(ws, row, "공모사항")
    row += 1

    securities = offering.get("securities", [{}])
    sec = securities[0] if securities else {}

    offering_items = [
        ("공모주식수", sec.get("count")),
        ("액면가", sec.get("face_value")),
        ("공모가(확정)", crawler.get("confirmed_price", "")),
        ("공모가(밴드)", offering.get("crawler_offering_price_range", "")),
        ("공모총액", sec.get("total_amount")),
        ("공모방법", sec.get("method", "")),
    ]
    for label, val in offering_items:
        ws.cell(row=row, column=1, value=label).font = HEADER_FONT
        cell = ws.cell(row=row, column=2, value=val)
        if isinstance(val, (int, float)):
            cell.number_format = NUM_FORMAT_KRW
        row += 1

    # 주관사
    underwriters = offering.get("underwriters", [])
    if underwriters:
        row += 1
        ws.cell(row=row, column=1, value="주관사").font = HEADER_FONT
        ws.cell(row=row, column=2, value="인수수량")
        ws.cell(row=row, column=3, value="인수금액")
        row += 1
        for uw in underwriters:
            ws.cell(row=row, column=1, value=uw.get("name", ""))
            ws.cell(row=row, column=2, value=uw.get("count"))
            ws.cell(row=row, column=3, value=uw.get("amount"))
            row += 1

    row += 1

    # =====================================================================
    # 4. 수요예측 결과
    # =====================================================================
    if crawler:
        _write_section_header(ws, row, "수요예측 결과")
        row += 1
        demand_items = [
            ("기관경쟁률", crawler.get("institutional_competition", "")),
            ("의무보유확약", crawler.get("lockup_commitment", "")),
            ("기관배정비율", crawler.get("institutional_allocation", "")),
            ("일반배정비율", crawler.get("retail_allocation", "")),
        ]
        for label, val in demand_items:
            if val:
                ws.cell(row=row, column=1, value=label).font = HEADER_FONT
                ws.cell(row=row, column=2, value=val)
                row += 1
        row += 1

    # =====================================================================
    # 5. 유통가능주식수
    # =====================================================================
    lockup = data.get("lockup_schedule", [])
    if lockup:
        _write_section_header(ws, row, "유통가능주식수")
        row += 1
        headers = ["의무보유기간", "주식수", "비율", "누적비율"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h).font = HEADER_FONT
        row += 1
        for item in lockup:
            ws.cell(row=row, column=1, value=item.get("period", ""))
            ws.cell(row=row, column=2, value=item.get("shares"))
            ratio_cell = ws.cell(row=row, column=3, value=item.get("ratio"))
            cum_cell = ws.cell(row=row, column=4, value=item.get("cumulative_ratio"))
            if isinstance(item.get("ratio"), (int, float)):
                ratio_cell.number_format = NUM_FORMAT_PCT
            if isinstance(item.get("cumulative_ratio"), (int, float)):
                cum_cell.number_format = NUM_FORMAT_PCT
            row += 1
        row += 1

    # =====================================================================
    # 6. 재무제표
    # =====================================================================
    financials = data.get("financials", [])
    if financials:
        _write_section_header(ws, row, "Financial Highlights")
        row += 1
        fin_headers = ["구분"] + [f.get("year", "") for f in financials]
        for i, h in enumerate(fin_headers, 1):
            ws.cell(row=row, column=i, value=h).font = HEADER_FONT
        row += 1

        for field, label in [
            ("total_assets", "자산총계"),
            ("total_liabilities", "부채총계"),
            ("total_equity", "자본총계"),
            ("revenue", "매출액"),
            ("operating_income", "영업이익"),
            ("net_income", "당기순이익"),
        ]:
            ws.cell(row=row, column=1, value=label).font = HEADER_FONT
            for i, f in enumerate(financials, 2):
                cell = ws.cell(row=row, column=i, value=f.get(field))
                cell.number_format = NUM_FORMAT_KRW
            row += 1

        # YoY 성장률
        ws.cell(row=row, column=1, value="매출 YoY").font = HEADER_FONT
        for i, f in enumerate(financials, 2):
            yoy = f.get("revenue_yoy")
            if yoy is not None:
                cell = ws.cell(row=row, column=i, value=yoy)
                cell.number_format = NUM_FORMAT_PCT
        row += 1
        row += 1

    # =====================================================================
    # 7. Valuation
    # =====================================================================
    valuation = data.get("valuation", {})
    if valuation:
        _write_section_header(ws, row, "Valuation")
        row += 1

        val_items = [
            ("밸류에이션 방법", valuation.get("valuation_method", "")),
            ("기준 지표", valuation.get("base_metric", "")),
            ("기준 값", valuation.get("base_value")),
            ("할인율", valuation.get("discount_rate")),
            ("적용 배수", valuation.get("applied_multiple")),
            ("주당 평가가액", valuation.get("per_share_value")),
        ]
        for label, val in val_items:
            ws.cell(row=row, column=1, value=label).font = HEADER_FONT
            ws.cell(row=row, column=2, value=val)
            row += 1

        # 평균 PER
        avg_per = valuation.get("average_peer_per")
        if avg_per:
            ws.cell(row=row, column=1, value="비교회사 평균 PER").font = HEADER_FONT
            ws.cell(row=row, column=2, value=avg_per)
            row += 1

        # Peer Group
        peers = valuation.get("peers", [])
        if peers:
            row += 1
            ws.cell(row=row, column=1, value="Peer Group").font = HEADER_FONT
            row += 1
            peer_headers = ["회사", "거래소", "매출액", "당기순이익", "시가총액", "기준주가", "PER", "EV/EBITDA"]
            for i, h in enumerate(peer_headers, 1):
                ws.cell(row=row, column=i, value=h).font = HEADER_FONT
            row += 1
            for peer in peers:
                ws.cell(row=row, column=1, value=peer.get("name", ""))
                ws.cell(row=row, column=2, value=peer.get("market", ""))
                c_rev = ws.cell(row=row, column=3, value=peer.get("revenue"))
                c_ni = ws.cell(row=row, column=4, value=peer.get("net_income"))
                c_mc = ws.cell(row=row, column=5, value=peer.get("market_cap"))
                c_sp = ws.cell(row=row, column=6, value=peer.get("share_price"))
                ws.cell(row=row, column=7, value=peer.get("per"))
                ws.cell(row=row, column=8, value=peer.get("ev_ebitda"))
                for c in [c_rev, c_ni, c_mc, c_sp]:
                    if isinstance(c.value, (int, float)):
                        c.number_format = NUM_FORMAT_KRW
                row += 1
        row += 1

    # =====================================================================
    # 8. 종합의견 (AI 분석 요약)
    # =====================================================================
    analysis = data.get("analysis_report", "")
    if analysis:
        _write_section_header(ws, row, "종합의견")
        row += 1
        # 분석 리포트에서 종합 의견 섹션만 추출
        opinion_section = _extract_opinion(analysis)
        ws.cell(row=row, column=1, value=opinion_section).alignment = WRAP_ALIGNMENT
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 10, end_column=13)
        ws.row_dimensions[row].height = 200
        row += 12

    # 테두리 적용
    for r in range(1, row):
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER

    # 저장
    today = datetime.now().strftime("%Y%m%d")
    filename = f"{today}_{company_name}_리서치.xlsx"
    filepath = REPORTS_DIR / filename
    wb.save(filepath)
    print(f"[엑셀] 저장 완료: {filepath}")
    return filepath


def _extract_opinion(report: str) -> str:
    """분석 리포트에서 종합의견 + 체크포인트를 추출한다."""
    lines = report.split("\n")
    capture = False
    result = []

    for line in lines:
        if "종합 의견" in line or "종합의견" in line or "핵심 체크" in line:
            capture = True
        if capture:
            result.append(line)

    return "\n".join(result) if result else report[-2000:]
